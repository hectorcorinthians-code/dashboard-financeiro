
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

DATA_FILE = Path(__file__).resolve().parents[1] / "data" / "controle_baumevie_completo_estoque.xlsx"


@dataclass
class SheetConfig:
    name: str
    header_row: int
    start_row: int
    end_col: int | None = None


SHEET_CONFIG: dict[str, SheetConfig] = {
    "Produtos": SheetConfig("Produtos", header_row=3, start_row=4, end_col=12),
    "Vendas diárias": SheetConfig("Vendas diárias", header_row=3, start_row=4, end_col=10),
    "Financeiro": SheetConfig("Financeiro", header_row=3, start_row=4, end_col=8),
    "Resumo": SheetConfig("Resumo", header_row=3, start_row=4, end_col=8),
    "Vendas": SheetConfig("Vendas", header_row=1, start_row=2, end_col=4),
    "Estoque inicio Loja": SheetConfig("Estoque inicio Loja", header_row=1, start_row=2, end_col=23),
}


def workbook_path() -> Path:
    return DATA_FILE


def list_sheets() -> list[str]:
    return list(SHEET_CONFIG.keys())


def _load_workbook(data_only: bool = False):
    return openpyxl.load_workbook(DATA_FILE, data_only=data_only)


def _header_to_text(value: Any, fallback: str) -> str:
    if value is None or str(value).strip() == "":
        return fallback
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    return text if text else fallback


def _make_unique(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for idx, header in enumerate(headers, start=1):
        base = header or f"Coluna {idx}"
        count = seen.get(base, 0)
        if count == 0:
            unique.append(base)
        else:
            unique.append(f"{base} ({count + 1})")
        seen[base] = count + 1
    return unique


def _get_headers(ws, cfg: SheetConfig) -> list[str]:
    end_col = cfg.end_col or ws.max_column
    headers = []
    for c in range(1, end_col + 1):
        fallback = "Produto" if cfg.name == "Estoque inicio Loja" and c == 1 else f"Coluna {c}"
        headers.append(_header_to_text(ws.cell(cfg.header_row, c).value, fallback))
    return _make_unique(headers)


def _is_blank_row(values: list[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _series_or_zeros(df: pd.DataFrame, column: str) -> pd.Series:
    if column in df.columns:
        return pd.to_numeric(df[column], errors="coerce").fillna(0)
    return pd.Series(0, index=df.index, dtype="float64")


def get_sheet_dataframe(sheet_name: str) -> pd.DataFrame:
    cfg = SHEET_CONFIG[sheet_name]
    wb = _load_workbook(data_only=True)
    ws = wb[sheet_name]
    end_col = cfg.end_col or ws.max_column
    headers = _get_headers(ws, cfg)

    rows = []
    for r in range(cfg.start_row, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, end_col + 1)]
        if _is_blank_row(values):
            continue
        rows.append(values)

    df = pd.DataFrame(rows, columns=headers)

    for column in df.columns:
        col_lower = str(column).lower()
        if col_lower.startswith("data") or str(column) == "RECEBIMENTO":
            df[column] = pd.to_datetime(df[column], errors="coerce", dayfirst=True)

    return df


def save_sheet_dataframe(sheet_name: str, df: pd.DataFrame) -> None:
    cfg = SHEET_CONFIG[sheet_name]
    wb = _load_workbook(data_only=False)
    ws = wb[sheet_name]
    end_col = cfg.end_col or ws.max_column
    headers = _get_headers(ws, cfg)

    existing_row_count = max(ws.max_row - cfg.start_row + 1, 0)
    new_row_count = len(df)
    total_rows = max(existing_row_count, new_row_count)

    for row_offset in range(total_rows):
        excel_row = cfg.start_row + row_offset
        has_df_row = row_offset < new_row_count

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(excel_row, col_idx)

            if not has_df_row:
                cell.value = None
                continue

            value = df.iloc[row_offset][header]
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif hasattr(value, "item"):
                value = value.item()

            cell.value = value

    wb.save(DATA_FILE)


def get_sheet_summary(sheet_name: str) -> dict[str, Any]:
    df = get_sheet_dataframe(sheet_name).copy()

    numeric_sums = []
    total_numeric_sum = 0.0

    for col in df.columns:
        series = pd.to_numeric(df[col], errors="coerce")
        if series.notna().any():
            col_sum = float(series.fillna(0).sum())
            total_numeric_sum += col_sum
            numeric_sums.append({"Coluna": col, "Soma": col_sum})

    top_numeric_sums = (
        pd.DataFrame(numeric_sums).sort_values("Soma", ascending=False).head(10)
        if numeric_sums
        else pd.DataFrame(columns=["Coluna", "Soma"])
    )

    return {
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "numeric_columns": int(len(numeric_sums)),
        "numeric_sum": total_numeric_sum,
        "top_numeric_sums": top_numeric_sums,
    }


def build_dashboard_metrics() -> dict[str, Any]:
    produtos = get_sheet_dataframe("Produtos").copy()
    vendas = get_sheet_dataframe("Vendas diárias").copy()
    financeiro = get_sheet_dataframe("Financeiro").copy()
    recebimentos = get_sheet_dataframe("Vendas").copy()

    for col in ["Estoque Inicial", "AMOSTRA", "Preço de Venda (R$)", "Custo Unitário (R$)", "Estoque Mínimo"]:
        if col in produtos.columns:
            produtos[col] = pd.to_numeric(produtos[col], errors="coerce")

    for col in ["Quantidade", "Preço Unitário (R$)", "Valor Total (R$)", "Custo Total (R$)", "Lucro Bruto (R$)"]:
        if col in vendas.columns:
            vendas[col] = pd.to_numeric(vendas[col], errors="coerce")

    for col in ["Valor", "Saldo conta"]:
        if col in recebimentos.columns:
            recebimentos[col] = pd.to_numeric(recebimentos[col], errors="coerce")

    product_cost_map = {}
    product_price_map = {}
    if {"Produto", "Custo Unitário (R$)"}.issubset(produtos.columns):
        product_cost_map = produtos.set_index("Produto")["Custo Unitário (R$)"].to_dict()
    if {"Produto", "Preço de Venda (R$)"}.issubset(produtos.columns):
        product_price_map = produtos.set_index("Produto")["Preço de Venda (R$)"].to_dict()

    if "Produto" in vendas.columns and "Preço Unitário (R$)" in vendas.columns:
        vendas["Preço Unitário (R$)"] = vendas.apply(
            lambda row: row["Preço Unitário (R$)"] if pd.notna(row["Preço Unitário (R$)"]) else product_price_map.get(row.get("Produto")),
            axis=1,
        )

    vendas["Quantidade"] = _series_or_zeros(vendas, "Quantidade")
    vendas["Preço Unitário (R$)"] = _series_or_zeros(vendas, "Preço Unitário (R$)")
    vendas["Valor Total (R$)"] = _series_or_zeros(vendas, "Valor Total (R$)")
    vendas["Custo Total (R$)"] = _series_or_zeros(vendas, "Custo Total (R$)")
    vendas["Lucro Bruto (R$)"] = _series_or_zeros(vendas, "Lucro Bruto (R$)")

    vendas["Valor Calculado (R$)"] = vendas["Quantidade"] * vendas["Preço Unitário (R$)"]
    mask_valor_zero = vendas["Valor Total (R$)"].eq(0)
    vendas.loc[mask_valor_zero, "Valor Total (R$)"] = vendas.loc[mask_valor_zero, "Valor Calculado (R$)"]

    if "Produto" in vendas.columns:
        vendas["Custo Lookup (R$)"] = vendas["Produto"].map(product_cost_map).fillna(0)
    else:
        vendas["Custo Lookup (R$)"] = 0

    mask_custo_zero = vendas["Custo Total (R$)"].eq(0)
    vendas.loc[mask_custo_zero, "Custo Total (R$)"] = (
        vendas.loc[mask_custo_zero, "Quantidade"] * vendas.loc[mask_custo_zero, "Custo Lookup (R$)"]
    )

    if "Tipo" in vendas.columns:
        tipo_normalizado = vendas["Tipo"].astype(str).str.strip().str.lower()
        is_sale = tipo_normalizado.eq("venda")
        is_entry = tipo_normalizado.eq("entrada")
    else:
        is_sale = pd.Series(True, index=vendas.index)
        is_entry = pd.Series(False, index=vendas.index)

    mask_lucro_zero = vendas["Lucro Bruto (R$)"].eq(0)
    vendas.loc[mask_lucro_zero, "Lucro Bruto (R$)"] = vendas["Valor Total (R$)"] - vendas["Custo Total (R$)"]
    vendas.loc[~is_sale, "Lucro Bruto (R$)"] = 0

    entradas_por_produto = (
        vendas[is_entry].groupby("Produto")["Quantidade"].sum()
        if "Produto" in vendas.columns
        else pd.Series(dtype=float)
    )
    vendas_por_produto = (
        vendas[is_sale].groupby("Produto")["Quantidade"].sum()
        if "Produto" in vendas.columns
        else pd.Series(dtype=float)
    )

    produtos["Entradas Calc"] = produtos["Produto"].map(entradas_por_produto).fillna(0) if "Produto" in produtos.columns else 0
    produtos["Vendas Calc"] = produtos["Produto"].map(vendas_por_produto).fillna(0) if "Produto" in produtos.columns else 0
    produtos["AMOSTRA"] = _series_or_zeros(produtos, "AMOSTRA")
    produtos["Estoque Inicial"] = _series_or_zeros(produtos, "Estoque Inicial")
    produtos["Preço de Venda (R$)"] = _series_or_zeros(produtos, "Preço de Venda (R$)")
    produtos["Custo Unitário (R$)"] = _series_or_zeros(produtos, "Custo Unitário (R$)")
    produtos["Estoque Mínimo"] = _series_or_zeros(produtos, "Estoque Mínimo")

    produtos["Estoque Atual Calc"] = (
        produtos["Estoque Inicial"] + produtos["Entradas Calc"] - produtos["Vendas Calc"] - produtos["AMOSTRA"]
    )
    produtos["Valor em Estoque Calc"] = produtos["Estoque Atual Calc"] * produtos["Preço de Venda (R$)"]

    sales_only = vendas[is_sale].copy()
    entries_only = vendas[is_entry].copy()

    faturamento_total = float(sales_only["Valor Total (R$)"].sum())
    lucro_total = float(sales_only["Lucro Bruto (R$)"].sum())
    itens_vendidos = float(sales_only["Quantidade"].sum())
    pedidos = int((sales_only["Quantidade"] > 0).sum())
    ticket_medio = faturamento_total / pedidos if pedidos else 0.0
    valor_estoque = float(produtos["Valor em Estoque Calc"].sum())
    custo_estoque = float((produtos["Estoque Atual Calc"] * produtos["Custo Unitário (R$)"]).sum())
    total_entradas = float(entries_only["Quantidade"].sum())
    total_recebido = float(recebimentos["Valor"].fillna(0).sum()) if "Valor" in recebimentos.columns else 0.0
    saldo_conta = (
        float(recebimentos["Saldo conta"].dropna().iloc[-1])
        if "Saldo conta" in recebimentos.columns and not recebimentos["Saldo conta"].dropna().empty
        else 0.0
    )

    produtos_repor = produtos[produtos["Estoque Atual Calc"] <= produtos["Estoque Mínimo"]].copy()

    sales_by_day = pd.DataFrame()
    if {"Data", "Valor Total (R$)"}.issubset(sales_only.columns):
        sales_by_day = (
            sales_only.dropna(subset=["Data"])
            .groupby(pd.Grouper(key="Data", freq="D"))["Valor Total (R$)"]
            .sum()
            .reset_index()
            .rename(columns={"Valor Total (R$)": "Faturamento"})
        )

    sales_by_channel = pd.DataFrame()
    if {"Canal de Venda", "Valor Total (R$)"}.issubset(sales_only.columns):
        sales_by_channel = (
            sales_only.groupby("Canal de Venda", dropna=False)["Valor Total (R$)"]
            .sum()
            .reset_index()
            .sort_values("Valor Total (R$)", ascending=False)
        )

    profit_by_channel = pd.DataFrame()
    if {"Canal de Venda", "Lucro Bruto (R$)"}.issubset(sales_only.columns):
        profit_by_channel = (
            sales_only.groupby("Canal de Venda", dropna=False)["Lucro Bruto (R$)"]
            .sum()
            .reset_index()
            .sort_values("Lucro Bruto (R$)", ascending=False)
        )

    canal_destaque = str(sales_by_channel.iloc[0]["Canal de Venda"]) if not sales_by_channel.empty else None

    sales_by_product = pd.DataFrame()
    if {"Produto", "Quantidade", "Valor Total (R$)"}.issubset(sales_only.columns):
        sales_by_product = (
            sales_only.groupby("Produto", dropna=False)[["Quantidade", "Valor Total (R$)"]]
            .sum()
            .reset_index()
            .sort_values("Valor Total (R$)", ascending=False)
        )

    inventory_by_category = pd.DataFrame()
    if {"Categoria", "Estoque Atual Calc", "Valor em Estoque Calc"}.issubset(produtos.columns):
        inventory_by_category = (
            produtos.groupby("Categoria", dropna=False)[["Estoque Atual Calc", "Valor em Estoque Calc"]]
            .sum()
            .reset_index()
            .sort_values("Valor em Estoque Calc", ascending=False)
            .rename(columns={"Estoque Atual Calc": "Estoque Atual", "Valor em Estoque Calc": "Valor em Estoque varejo(R$)"})
        )

    return {
        "faturamento_total": faturamento_total,
        "lucro_total": lucro_total,
        "itens_vendidos": itens_vendidos,
        "ticket_medio": ticket_medio,
        "valor_estoque": valor_estoque,
        "custo_estoque": custo_estoque,
        "total_entradas": total_entradas,
        "total_recebido": total_recebido,
        "saldo_conta": saldo_conta,
        "canal_destaque": canal_destaque,
        "produtos_repor": produtos_repor,
        "sales_by_day": sales_by_day,
        "sales_by_channel": sales_by_channel,
        "profit_by_channel": profit_by_channel,
        "sales_by_product": sales_by_product,
        "inventory_by_category": inventory_by_category,
        "produtos": produtos,
        "vendas": vendas,
        "financeiro": financeiro,
    }
